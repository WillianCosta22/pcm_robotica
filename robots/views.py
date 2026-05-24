"""Views do app Robots"""
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.db.models import Q
from django.http import HttpResponse
from io import BytesIO

from .models import Robot
from .forms import RobotForm, RobotFilterForm


class AdminRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser


class RobotListView(LoginRequiredMixin, ListView):
    model = Robot
    template_name = 'robots/robot_list.html'
    context_object_name = 'robots'
    paginate_by = 15

    def get_queryset(self):
        queryset = Robot.objects.all()
        form = RobotFilterForm(self.request.GET)
        if form.is_valid():
            busca = form.cleaned_data.get('busca')
            status = form.cleaned_data.get('status')
            localizacao = form.cleaned_data.get('localizacao')
            if busca:
                queryset = queryset.filter(
                    Q(nome__icontains=busca) | Q(modelo__icontains=busca) |
                    Q(numero_serie__icontains=busca) | Q(patrimonio__icontains=busca) |
                    Q(fabricante__icontains=busca) | Q(alocacao__icontains=busca)
                )
            if status:
                queryset = queryset.filter(status_operacional=status)
            if localizacao:
                queryset = queryset.filter(localizacao__icontains=localizacao)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = RobotFilterForm(self.request.GET)
        context['total'] = Robot.objects.count()
        return context


class RobotDetailView(LoginRequiredMixin, DetailView):
    model = Robot
    template_name = 'robots/robot_detail.html'
    context_object_name = 'robot'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        robot = self.get_object()
        context['manutencoes'] = robot.manutencao_set.order_by('-data_abertura')[:10]
        return context


class RobotCreateView(LoginRequiredMixin, AdminRequiredMixin, CreateView):
    model = Robot
    form_class = RobotForm
    template_name = 'robots/robot_form.html'
    success_url = reverse_lazy('robots:list')

    def form_valid(self, form):
        messages.success(self.request, '✅ Robô cadastrado com sucesso!')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Cadastrar Robô'
        context['btn_label'] = 'Cadastrar'
        return context


class RobotUpdateView(LoginRequiredMixin, AdminRequiredMixin, UpdateView):
    model = Robot
    form_class = RobotForm
    template_name = 'robots/robot_form.html'
    success_url = reverse_lazy('robots:list')

    def form_valid(self, form):
        messages.success(self.request, '✅ Robô atualizado com sucesso!')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Editar Robô: {self.object.nome}'
        context['btn_label'] = 'Salvar Alterações'
        return context


class RobotDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = Robot
    template_name = 'robots/robot_confirm_delete.html'
    success_url = reverse_lazy('robots:list')

    def form_valid(self, form):
        messages.success(self.request, '🗑️ Robô excluído com sucesso!')
        return super().form_valid(form)


def exportar_robots_excel(request):
    """Exporta lista de robôs em Excel formatado"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Gestão de Ativos'

    # Header principal
    ws.merge_cells('A1:L1')
    ws['A1'] = 'Planilha De Gestão de Ativos - Robótica'
    ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', start_color='1F2D40')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Sub-cabeçalho
    headers = ['ATIVO', 'PATRIMÔNIO', 'DENOMINAÇÃO', 'TIPO DE MNT', 'GRUPO',
               'ALOCAÇÃO', 'STATUS', 'MODELO', 'FABRICANTE', 'LOCALIZAÇÃO', 'OBSERVAÇÕES']
    header_row = 2
    header_fill = PatternFill('solid', start_color='2E4057')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF', size=9)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[header_row].height = 22

    # Data rows
    STATUS_LABELS = {
        'operacional': 'Operacional', 'em_manutencao': 'Em manutenção',
        'parado': 'Parado', 'em_inspecao': 'Em inspeção', 'inativo': 'Inativo',
    }
    STATUS_COLORS = {
        'operacional': '198754', 'em_manutencao': 'FFC107',
        'parado': 'DC3545', 'em_inspecao': '0DCAF0', 'inativo': '6C757D',
    }

    row_fills = ['F8F9FA', 'FFFFFF']
    for i, robot in enumerate(Robot.objects.all()):
        row = i + 3
        fill = PatternFill('solid', start_color=row_fills[i % 2])
        vals = [
            robot.numero_serie or '',
            robot.patrimonio,
            robot.nome,
            '',  # tipo mnt — vazio
            robot.grupo,
            robot.alocacao,
            STATUS_LABELS.get(robot.status_operacional, robot.status_operacional),
            robot.modelo,
            robot.fabricante,
            robot.localizacao,
            robot.observacoes,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical='center', wrap_text=True)

        # Color status cell
        status_cell = ws.cell(row=row, column=7)
        color = STATUS_COLORS.get(robot.status_operacional, '6C757D')
        status_cell.font = Font(bold=True, color=color, size=9)

    # Column widths
    widths = [14, 12, 40, 12, 12, 14, 14, 20, 18, 20, 30]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Freeze header rows
    ws.freeze_panes = 'A3'

    # Save
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="gestao_ativos_robotica.xlsx"'
    return response


import json
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt


@login_required
def preview_importacao_excel(request):
    """
    Recebe o arquivo Excel via POST (AJAX/form) e devolve JSON
    com as linhas lidas — sem salvar nada no banco ainda.
    Colunas esperadas (mesma ordem do exportar_robots_excel):
      A=ATIVO, B=PATRIMÔNIO, C=DENOMINAÇÃO, D=TIPO_MNT, E=GRUPO,
      F=ALOCAÇÃO, G=STATUS, H=MODELO, I=FABRICANTE, J=LOCALIZAÇÃO, K=OBSERVAÇÕES
    Também aceita a variante com linha de título na linha 1 e cabeçalho na linha 2.
    """
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponse(status=403)

    if request.method != 'POST':
        return HttpResponse(status=405)

    arquivo = request.FILES.get('arquivo')
    if not arquivo:
        return HttpResponse(json.dumps({'erro': 'Nenhum arquivo enviado.'}),
                            content_type='application/json', status=400)

    from openpyxl import load_workbook
    try:
        wb = load_workbook(arquivo, data_only=True)
        ws = wb.active

        # Detectar linha inicial de dados:
        # Se a linha 1 parece título (células mescladas ou texto longo) e
        # a linha 2 tem texto parecido com cabeçalho → dados começam na linha 3.
        # Caso contrário, linha 1 = cabeçalho → dados começam na linha 2.
        primeira_celula_r1 = str(ws.cell(1, 1).value or '').strip().upper()
        primeira_celula_r2 = str(ws.cell(2, 1).value or '').strip().upper()

        cabecalhos_conhecidos = {'ATIVO', 'PATRIMÔNIO', 'DENOMINAÇÃO', 'NOME', 'PATRIMONIO', 'DENOMINACAO'}
        if primeira_celula_r2 in cabecalhos_conhecidos or primeira_celula_r1 not in cabecalhos_conhecidos:
            # formato com título na linha 1
            start_row = 3
        else:
            start_row = 2

        STATUS_MAP = {
            'operacional': 'operacional',
            'em manutenção': 'em_manutencao', 'em manutencao': 'em_manutencao',
            'parado': 'parado', 'parada': 'parado',
            'em inspeção': 'em_inspecao', 'em inspecao': 'em_inspecao',
            'inativo': 'inativo', 'inativa': 'inativo',
            'inoperante': 'inativo',
        }

        linhas = []
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            # Pular linhas totalmente vazias
            if all(v is None or str(v).strip() == '' for v in row):
                continue

            def _s(v): return str(v).strip() if v is not None else ''

            numero_serie = _s(row[0]) if len(row) > 0 else ''
            patrimonio   = _s(row[1]) if len(row) > 1 else ''
            nome         = _s(row[2]) if len(row) > 2 else ''
            grupo        = _s(row[4]) if len(row) > 4 else ''
            alocacao     = _s(row[5]) if len(row) > 5 else ''
            status_raw   = _s(row[6]).lower() if len(row) > 6 else ''
            modelo       = _s(row[7]) if len(row) > 7 else ''
            fabricante   = _s(row[8]) if len(row) > 8 else ''
            localizacao  = _s(row[9]) if len(row) > 9 else ''
            observacoes  = _s(row[10]) if len(row) > 10 else ''

            if not nome and not numero_serie and not patrimonio:
                continue

            status = STATUS_MAP.get(status_raw, 'operacional')

            # Verificar se já existe no banco
            existe = False
            conflito = ''
            if numero_serie:
                existe = Robot.objects.filter(numero_serie=numero_serie).exists()
                if existe:
                    conflito = f'Ativo "{numero_serie}" já existe'
            if not existe and patrimonio:
                existe = Robot.objects.filter(patrimonio=patrimonio).exists()
                if existe:
                    conflito = f'Patrimônio "{patrimonio}" já existe'

            linhas.append({
                'numero_serie': numero_serie,
                'patrimonio':   patrimonio,
                'nome':         nome,
                'grupo':        grupo,
                'alocacao':     alocacao,
                'status':       status,
                'status_display': dict(Robot.STATUS_CHOICES).get(status, status),
                'modelo':       modelo,
                'fabricante':   fabricante,
                'localizacao':  localizacao,
                'observacoes':  observacoes,
                'existe':       existe,
                'conflito':     conflito,
            })

        return HttpResponse(json.dumps({'linhas': linhas, 'total': len(linhas)}),
                            content_type='application/json')

    except Exception as e:
        return HttpResponse(json.dumps({'erro': f'Erro ao ler arquivo: {str(e)}'}),
                            content_type='application/json', status=400)


@login_required
@require_POST
def confirmar_importacao_excel(request):
    """
    Recebe JSON com a lista de robôs selecionados e os salva no banco.
    Body: { "robots": [...], "sobrescrever": true/false }
    """
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponse(status=403)

    try:
        body = json.loads(request.body)
    except Exception:
        return HttpResponse(json.dumps({'erro': 'JSON inválido'}),
                            content_type='application/json', status=400)

    robots_data = body.get('robots', [])
    sobrescrever = body.get('sobrescrever', False)

    criados = 0
    atualizados = 0
    ignorados = 0
    erros = []

    for rd in robots_data:
        numero_serie = rd.get('numero_serie') or None
        patrimonio   = rd.get('patrimonio', '')
        nome         = rd.get('nome', '').strip()

        if not nome:
            ignorados += 1
            continue

        # Tentar encontrar registro existente
        existing = None
        if numero_serie:
            existing = Robot.objects.filter(numero_serie=numero_serie).first()
        if not existing and patrimonio:
            existing = Robot.objects.filter(patrimonio=patrimonio).first()

        campos = {
            'nome':           nome,
            'patrimonio':     patrimonio,
            'modelo':         rd.get('modelo', ''),
            'fabricante':     rd.get('fabricante', ''),
            'grupo':          rd.get('grupo', ''),
            'alocacao':       rd.get('alocacao', ''),
            'localizacao':    rd.get('localizacao', ''),
            'observacoes':    rd.get('observacoes', ''),
            'status_operacional': rd.get('status', 'operacional'),
        }
        if numero_serie:
            campos['numero_serie'] = numero_serie

        try:
            if existing:
                if sobrescrever:
                    for k, v in campos.items():
                        setattr(existing, k, v)
                    existing.save()
                    atualizados += 1
                else:
                    ignorados += 1
            else:
                Robot.objects.create(**campos)
                criados += 1
        except Exception as e:
            erros.append(f'{nome}: {str(e)}')

    return HttpResponse(json.dumps({
        'criados': criados,
        'atualizados': atualizados,
        'ignorados': ignorados,
        'erros': erros,
    }), content_type='application/json')
